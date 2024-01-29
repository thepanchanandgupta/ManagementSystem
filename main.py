import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl

combo_list = ["IT", "Admin", "HR", "Others"]
cols = ("Name", "Age", "Department", "Notice Period", "Experience Level")


def toggle_mode():
    if mode_switch.instate(["selected"]):
        style.theme_use("forest-light")
    else:
        style.theme_use("forest-dark")


def load_data():
    path = "/Users/panchanandgupta/anaconda3/lib/python3.11/site-packages/jupyterlab_server/test_data/workspaces/workSpace/PycharmProjects/ManagementSystem/employeeDB.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)

    for col_name in list_values[0]:
        treeView.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeView.insert('', tk.END, values=value_tuple)


def insert_row():
    name = name_entry.get()
    age = age_spinbox.get()
    department = department_combobox.get()
    employment_status = "Yes" if a.get() else "No"
    # ( "No" if not a.get() else messagebox.showwarning("Warning", "Field Empty"))
    experience_level = "Fresher" if b.get() == "Fresher" else "Experienced"
    # ( "Experienced" if b.get() == "Experienced" else messagebox.showwarning("Warning", "Field Empty"))

    # Insert row into Excel Sheet
    path = "/Users/panchanandgupta/anaconda3/lib/python3.11/site-packages/jupyterlab_server/test_data/workspaces/workSpace/PycharmProjects/ManagementSystem/employeeDB.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    if name != "Name" and age != "Age" and department != "--Select Department" and experience_level != "ok" and employment_status:
        row_values = [name, age, department, employment_status, experience_level]
        sheet.append(row_values)
        workbook.save(path)

        # Insert row into TreeView
        treeView.insert('', tk.END, values=row_values)

        messagebox.showinfo("Status", "Data Submitted")

        # Set to Default
        name_entry.delete(0, "end")
        name_entry.insert(0, "Name")
        age_spinbox.delete(0, "end")
        age_spinbox.insert(0, "Age")
        department_combobox.set("--Select Department--")
        noticePeriod_checkButton.state(["!selected"])
        b.set(None)

    else:
        messagebox.showwarning("Warning", "Field(s) Empty.")

    workbook.close()


def read_data():
    read_query = read_entry.get().lower()
    if read_query == "":
        messagebox.showwarning("Warning", "No row selected.")
        return

    for item in treeView.get_children():
        treeView.delete(item)

    path = "/Users/panchanandgupta/anaconda3/lib/python3.11/site-packages/jupyterlab_server/test_data/workspaces/workSpace/PycharmProjects/ManagementSystem/employeeDB.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    read_column = cols

    for row in sheet.iter_rows(values_only=True):
        row_data = [str(val).lower() if val is not None else "" for val in row]

        if any(read_query in col_data for col_data in row_data for col in read_column):
            treeView.insert("", tk.END, values=row)

    workbook.close()


def update_row():
    selected_item = treeView.focus()
    if not selected_item:
        messagebox.showwarning("Warning", "Please select a row to update.")
        return

    current_values = treeView.item(selected_item, 'values')

    path = "/Users/panchanandgupta/anaconda3/lib/python3.11/site-packages/jupyterlab_server/test_data/workspaces/workSpace/PycharmProjects/ManagementSystem/employeeDB.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    selected_index = treeView.index(selected_item)

    # Get updated values from the user
    updated_name = name_entry.get() if name_entry.get() != "Name" else current_values[0]
    updated_age = age_spinbox.get() if age_spinbox.get() != "Select Age of Employee" else current_values[1]
    updated_department = department_combobox.get() if department_combobox.get() != "--Select Department--" else current_values[2]
    updated_employment_status = "Yes" if a.get() else "No" if current_values[3] == "Yes" else current_values[3]
    updated_experience_level = "Fresher" if b.get() == "Fresher" else ("Experienced" if b.get() == "Experienced" else current_values[4])

    sheet.cell(row=selected_index + 2, column=1, value=updated_name if updated_name else "")
    sheet.cell(row=selected_index + 2, column=2, value=updated_age if updated_age else "")
    sheet.cell(row=selected_index + 2, column=3, value=updated_department)
    sheet.cell(row=selected_index + 2, column=4, value=updated_employment_status)
    sheet.cell(row=selected_index + 2, column=5, value=updated_experience_level)

    workbook.save(path)

    treeView.item(selected_item, values=(updated_name, updated_age, updated_department, updated_employment_status, updated_experience_level))

    # Set to Default
    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Select Age of Employee")
    department_combobox.set("--Select Department--")
    noticePeriod_checkButton.state(["!selected"])
    b.set(None)

    workbook.close()


def delete_row():
    selected_item = treeView.focus()
    if not selected_item:
        messagebox.showwarning("Warning", "No row selected.")
        return

    # Getting the values from the Selected Row
    values = treeView.item(selected_item, 'values')
    print(values)

    # Open the workbook and select the active sheet
    path = "/Users/panchanandgupta/anaconda3/lib/python3.11/site-packages/jupyterlab_server/test_data/workspaces/workSpace/PycharmProjects/ManagementSystem/employeeDB.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # Find the row index based on the values
    selected_index = treeView.index(selected_item)
    sheet.delete_rows(selected_index + 2)
    treeView.delete(selected_item)

    workbook.save(path)

    # Set to Default
    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Age")
    department_combobox.set("--Select Department--")
    noticePeriod_checkButton.state(["!selected"])
    b.set(None)

    workbook.close()


root = tk.Tk()
root.title("Employee Management System")

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insert Row")
widgets_frame.grid(row=0, column=0, sticky="w", padx=20, pady=15)

# TextBox

name_entry = ttk.Entry(widgets_frame)
name_entry.insert(0, "Name")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=0, column=0, sticky="ew", padx=5, pady=10)

# Spinbox

age_spinbox = ttk.Spinbox(widgets_frame, from_=18, to=65)
age_spinbox.grid(row=1, column=0, sticky="ew", padx=5, pady=(0, 5))
age_spinbox.set("Select Age of Employee")
age_spinbox.bind("<FocusIn>", lambda e: age_spinbox.delete('0', 'end'))

# ComboBox aka DropDownList
department_combobox = ttk.Combobox(widgets_frame, values=combo_list)
department_combobox.set("--Select Department--")
department_combobox.grid(row=2, column=0, sticky="ew", padx=5, pady=6)

# CheckBox

a = tk.BooleanVar()
noticePeriod_checkButton = ttk.Checkbutton(widgets_frame, text="On Notice Period", variable=a)
noticePeriod_checkButton.grid(row=3, column=0, sticky="nsew", padx=5, pady=(0, 5))

# RadioButton

b = tk.StringVar()
fresher_radiobutton = ttk.Radiobutton(widgets_frame, text="Fresher", variable=b, value="Fresher")
experienced_radiobutton = ttk.Radiobutton(widgets_frame, text="Experienced", variable=b, value="Experienced")
fresher_radiobutton.grid(row=4, column=0, sticky="w", padx=5, pady=(0, 5))
experienced_radiobutton.grid(row=4, column=0, sticky="e", padx=5, pady=(0, 5))

# Separator

separator = ttk.Separator(widgets_frame)
separator.grid(row=5, column=0, sticky="nsew", padx=(10, 5), pady=10)

# Button

insert_button = ttk.Button(widgets_frame, text="Create", command=insert_row)
insert_button.grid(row=6, column=0, sticky="nsew", padx=10, pady=(0, 5))

read_entry = ttk.Entry(widgets_frame)
read_entry.insert(0, "Enter Name, Dept., Notice Period or Experience Level")
read_entry.bind("<FocusIn>", lambda e: read_entry.delete('0', 'end'))
read_entry.grid(row=7, column=0, sticky="nsew", padx=10, pady=(0, 5))

read_button = ttk.Button(widgets_frame, text="Search", command=read_data)
read_button.grid(row=8, column=0, columnspan=3, sticky="nsew", padx=10, pady=(0, 5))

update_button = ttk.Button(widgets_frame, text="Update", command=update_row)
update_button.grid(row=9, column=0, columnspan=3, sticky="nsew", padx=10, pady=(0, 5))

delete_button = ttk.Button(widgets_frame, text="Delete", command=delete_row)
delete_button.grid(row=10, column=0, columnspan=3, sticky="nsew", padx=10, pady=(0, 5))

# Separator

separator = ttk.Separator(widgets_frame)
separator.grid(row=11, column=0, sticky="ew", padx=(20, 10), pady=10)

# Toggle Switch

mode_switch = ttk.Checkbutton(widgets_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=12, column=0, sticky="nsew", padx=5, pady=10)

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

treeView = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=16,
                        selectmode='browse')
treeView.column("Name", width=100, anchor='center', stretch=tk.NO)
treeView.column("Age", width=50, anchor='center', stretch=tk.NO)
treeView.column("Department", width=100, anchor='center', stretch=tk.NO)
treeView.column("Notice Period", width=100, anchor='center', stretch=tk.NO)
treeView.column("Experience Level", width=100, anchor='center', stretch=tk.NO)
treeView.pack()
treeScroll.config(command=treeView.yview)

load_data()
root.mainloop()
